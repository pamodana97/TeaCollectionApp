
from openpyxl import load_workbook
import pandas as pd

class ExcelHandler:
    def __init__(self):
        self.workbook=None
        self.sheet=None

    def load(self,path):
        self.workbook=load_workbook(path)
        self.sheet=self.workbook.active
        return self.workbook,self.sheet

    def dataframe(self):
        data=self.sheet.values
        cols=next(data)
        df=pd.DataFrame(data,columns=cols)
        amt=df.columns[2:]
        df[amt]=df[amt].apply(pd.to_numeric,errors="coerce").round(2)
        return df.astype(object).where(pd.notna(df),"-")

    def find_customer_row(self,code):
        for r in range(2,self.sheet.max_row+1):
            if str(self.sheet.cell(r,1).value).strip()==str(code).strip():
                return r
        return None

    def find_date_column(self,day):
        for c in range(3,self.sheet.max_column+1):
            if self.sheet.cell(1,c).value==day:
                return c
        return None

    def update(self,code,day,amount):
        r=self.find_customer_row(code)
        c=self.find_date_column(day)
        if r is None or c is None:
            raise ValueError("Customer or date not found")
        self.sheet.cell(r,c).value=amount
        return r,c

    def save(self,path):
        self.workbook.save(path)

    def close(self):
        if self.workbook:
            self.workbook.close()

    def get_customer_data(self):
        """
        Returns customer dataframe and customer lookup dictionary.
        """

        df = self.dataframe()

        customer_df = df.iloc[:, 0:2].copy()

        customer_df.columns = [
            "Customer Code",
            "Customer Name"
        ]

        customer_df = customer_df.dropna(
            subset=["Customer Code"]
        )

        customer_df["Customer Code"] = (
            customer_df["Customer Code"].astype(str)
        )

        customer_dict = dict(
            zip(
                customer_df["Customer Code"],
                customer_df["Customer Name"]
            )
        )

        return customer_df, customer_dict
