import streamlit as st
import tempfile
import calendar
import pandas as pd
import io
import time
import os
import shutil
import textwrap
import streamlit.components.v1 as components

from datetime import datetime
from zoneinfo import ZoneInfo

# -------------------------------------------------------
# New Worksheet Template
# -------------------------------------------------------

TEMPLATE_PATH = os.path.join(
    os.path.dirname(__file__),
    "templates",
    "New Worksheet.xlsx"
)

from excel_handler import ExcelHandler
from session_manager import initialize
from ui_helpers import format_value, highlight

from auth import (
    login,
    logout
)

from audit import verify_audit_password

from database import (
    save_collection,
    delete_collection,
    get_last_collection
)

from workbook_storage import (
    upload_workbook,
    download_workbook,
    sync_workbook
)

from audit_db import (
    initialize_audit_db,
    add_audit_record,
    get_audit_records
)

DATE_EDIT_PASSWORD = "date123"
# -------------------------------------------------------
# Page Configuration
# -------------------------------------------------------

st.set_page_config(
    page_title="Tea Collection System",
    page_icon="🍃",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# -------------------------------------------------------
# Login Authentication
# -------------------------------------------------------

if not login():
    st.stop()

# -------------------------------------------------------
# Initialize Session
# -------------------------------------------------------

initialize()

# -------------------------------------------------------
# Initialize Audit Database
# -------------------------------------------------------

initialize_audit_db()

# -------------------------------------------------------
# Create Excel Handler
# -------------------------------------------------------

handler = ExcelHandler()

# -------------------------------------------------------
# Title
# -------------------------------------------------------

st.title("🍃 Tea Collection Entry System - WIJEWANTHA STORES")

# -------------------------------------------------------
# Current Sri Lanka Date
# -------------------------------------------------------

current_date = datetime.now(
    ZoneInfo("Asia/Colombo")
)

current_year = current_date.year
current_month = current_date.strftime("%B")

# -------------------------------------------------------
# Initialize Year / Month
# -------------------------------------------------------

if st.session_state["selected_year"] is None:
    st.session_state["selected_year"] = current_year

if st.session_state["selected_month"] is None:
    st.session_state["selected_month"] = current_month


# -------------------------------------------------------
# Load / Create Monthly Workbook
# -------------------------------------------------------

if st.session_state["temp_file"] is None:

    selected_year_for_workbook = (
        st.session_state["selected_year"]
    )

    selected_month_for_workbook = (
        st.session_state["selected_month"]
    )


    # ---------------------------------------------------
    # Try restoring workbook from Supabase
    # ---------------------------------------------------

    restored_file = download_workbook(
        selected_year_for_workbook,
        selected_month_for_workbook
    )


    if restored_file is not None:

        st.session_state["temp_file"] = (
            restored_file
        )

        st.session_state[
            "uploaded_filename"
        ] = (
            f"Tea_Collection_"
            f"{selected_year_for_workbook}_"
            f"{selected_month_for_workbook}.xlsx"
        )


    # ---------------------------------------------------
    # Workbook does NOT exist
    # ---------------------------------------------------

    else:

        st.info(
            f"No worksheet exists for "
            f"{selected_month_for_workbook} "
            f"{selected_year_for_workbook}."
        )


        if st.button(
            "➕ Add New Worksheet",
            type="primary",
            use_container_width=True,
            key="add_new_worksheet"
        ):

            try:

                # ---------------------------------------
                # Check template
                # ---------------------------------------

                if not os.path.exists(
                    TEMPLATE_PATH
                ):

                    st.error(
                        "New Worksheet template "
                        "could not be found."
                    )

                    st.stop()


                # ---------------------------------------
                # Create temporary workbook
                # ---------------------------------------

                new_workbook = (
                    tempfile.NamedTemporaryFile(
                        delete=False,
                        suffix=".xlsx"
                    )
                )

                new_workbook.close()


                # ---------------------------------------
                # Copy template
                # ---------------------------------------

                shutil.copyfile(
                    TEMPLATE_PATH,
                    new_workbook.name
                )


                # ---------------------------------------
                # Upload as monthly workbook
                # ---------------------------------------

                sync_workbook(
                    new_workbook.name,
                    selected_year_for_workbook,
                    selected_month_for_workbook
                )


                # ---------------------------------------
                # Use workbook immediately
                # ---------------------------------------

                st.session_state[
                    "temp_file"
                ] = new_workbook.name

                st.session_state[
                    "uploaded_filename"
                ] = (
                    f"Tea_Collection_"
                    f"{selected_year_for_workbook}_"
                    f"{selected_month_for_workbook}.xlsx"
                )


                st.session_state[
                    "new_worksheet_success"
                ] = True

                st.rerun()


            except Exception as e:

                st.error(
                    f"Unable to create worksheet: {e}"
                )


        # Do not execute Excel code until
        # a workbook exists
        st.stop()

# -------------------------------------------------------
# New Worksheet Success
# -------------------------------------------------------

if st.session_state.pop(
    "new_worksheet_success",
    False
):

    st.success(
        "✅ New worksheet created successfully."
    )

# -------------------------------------------------------
# Load workbook
# -------------------------------------------------------

workbook, sheet = handler.load(
    st.session_state["temp_file"]
)

df = handler.dataframe()

customer_df, customer_dict = handler.get_customer_data()


# Customer Code → Customer Name
code_to_name = dict(
    zip(
        customer_df["Customer Code"],
        customer_df["Customer Name"]
    )
)

# Customer Name → Customer Code
name_to_code = {
    str(name).strip(): code
    for name, code in zip(
        customer_df["Customer Name"],
        customer_df["Customer Code"]
    )
    if pd.notna(name)
}

# -------------------------------------------------------
# Entry Form
# -------------------------------------------------------

st.markdown("---")
st.subheader("Tea Collection Entry")


# -------------------------------------------------------
# Year / Month Access Control
# -------------------------------------------------------

date_col1, date_col2, date_col3 = st.columns(3)

month_names = [
    "January",
    "February",
    "March",
    "April",
    "May",
    "June",
    "July",
    "August",
    "September",
    "October",
    "November",
    "December"
]


# -------------------------------------------------------
# Year
# -------------------------------------------------------

with date_col1:

    if st.session_state.get(
        "date_fields_unlocked",
        False
    ):

        selected_year = st.number_input(
            "Year",
            min_value=2020,
            max_value=2100,
            value=int(
                st.session_state["selected_year"]
            ),
            step=1
        )

        st.session_state["selected_year"] = (
            selected_year
        )

    else:

        selected_year = st.number_input(
            "Year",
            value=int(
                st.session_state["selected_year"]
            ),
            disabled=True
        )


# -------------------------------------------------------
# Month
# -------------------------------------------------------

with date_col2:

    if st.session_state.get(
        "date_fields_unlocked",
        False
    ):

        current_month_index = month_names.index(
            st.session_state["selected_month"]
        )

        selected_month = st.selectbox(
            "Month",
            month_names,
            index=current_month_index
        )

        st.session_state["selected_month"] = (
            selected_month
        )

    else:

        selected_month = st.text_input(
            "Month",
            value=st.session_state[
                "selected_month"
            ],
            disabled=True
        )


# -------------------------------------------------------
# Day
# -------------------------------------------------------


with date_col3:

    # Convert month name to month number
    month_number = datetime.strptime(
        selected_month,
        "%B"
    ).month

    # Get number of days in selected month
    days_in_month = calendar.monthrange(
        selected_year,
        month_number
    )[1]

    # Display only valid days
    selected_day = st.selectbox(
        "Day",
        list(range(1, days_in_month + 1))
    )

# -------------------------------------------------------
# Unlock Year / Month
# -------------------------------------------------------

if not st.session_state.get(
    "date_fields_unlocked",
    False
):

    if st.button(
        "🔐 Change Year / Month",
        key="change_year_month"
    ):

        st.session_state[
            "show_date_password"
        ] = True


# -------------------------------------------------------
# Password Verification
# -------------------------------------------------------

if (
    st.session_state.get(
        "show_date_password",
        False
    )
    and not st.session_state.get(
        "date_fields_unlocked",
        False
    )
):

    date_password = st.text_input(
        "Enter password to change Year / Month",
        type="password",
        key="date_edit_password"
    )

    if st.button(
        "Unlock",
        key="unlock_date_fields"
    ):

        if date_password == DATE_EDIT_PASSWORD:

            st.session_state[
                "date_fields_unlocked"
            ] = True

            st.session_state[
                "show_date_password"
            ] = False

            st.rerun()

        else:

            st.error(
                "Incorrect password."
            )


# -------------------------------------------------------
# Lock Year / Month
# -------------------------------------------------------

if st.session_state.get(
    "date_fields_unlocked",
    False
):

    if st.button(
        "🔒 Lock Year / Month",
        key="lock_date_fields"
    ):

        st.session_state[
            "date_fields_unlocked"
        ] = False

        st.session_state[
            "show_date_password"
        ] = False

        # Force app to load workbook
        # for newly selected Year / Month
        st.session_state[
            "temp_file"
        ] = None

        st.session_state[
            "uploaded_filename"
        ] = None

        st.rerun()

# -------------------------------------------------------
# Customer Lists
# -------------------------------------------------------

customer_codes = (
    customer_df["Customer Code"].tolist()
)

customer_names = [
    str(name).strip()
    for name in customer_df["Customer Name"].tolist()
    if pd.notna(name)
]


# -------------------------------------------------------
# Initialize Customer Selection
# -------------------------------------------------------

if "customer_code_select" not in st.session_state:

    st.session_state["customer_code_select"] = (
        customer_codes[0]
    )


if "customer_name_select" not in st.session_state:

    first_code = st.session_state[
        "customer_code_select"
    ]

    st.session_state["customer_name_select"] = (
        code_to_name[first_code]
    )


# -------------------------------------------------------
# Customer Code Changed
# -------------------------------------------------------

def customer_code_changed():

    selected_code = st.session_state[
        "customer_code_select"
    ]

    corresponding_name = code_to_name.get(
        selected_code
    )

    if corresponding_name is not None:

        st.session_state[
            "customer_name_select"
        ] = corresponding_name

    # Move to Amount after selection
    st.session_state["focus_amount"] = True

# -------------------------------------------------------
# Customer Name Changed
# -------------------------------------------------------

def customer_name_changed():

    selected_name = st.session_state[
        "customer_name_select"
    ]

    corresponding_code = name_to_code.get(
        selected_name
    )

    if corresponding_code is not None:

        st.session_state[
            "customer_code_select"
        ] = corresponding_code

    # Move to Amount after selection
    st.session_state["focus_amount"] = True

# -------------------------------------------------------
# Customer Row
# -------------------------------------------------------

col1, col2 = st.columns([1, 2])


with col1:

    customer_code = st.selectbox(
        "Customer Code",
        customer_codes,
        key="customer_code_select",
        on_change=customer_code_changed
    )


with col2:

    customer_name = st.selectbox(
        "Customer Name",
        customer_names,
        key="customer_name_select",
        on_change=customer_name_changed
    )

# -------------------------------------------------------
# Customer Name - Select Entire Text on Click
# -------------------------------------------------------

components.html(
    """
    <script>
    const doc = window.parent.document;

    function setupCustomerNameSelectAll() {

        const labels = doc.querySelectorAll("label");

        labels.forEach((label) => {

            if (
                label.innerText.trim() !== "Customer Name"
            ) {
                return;
            }

            const container = label.parentElement;

            if (!container) {
                return;
            }

            const input = container.querySelector("input");

            if (!input) {
                return;
            }

            if (
                input.dataset.customerNameSelectAll === "true"
            ) {
                return;
            }

            input.dataset.customerNameSelectAll = "true";


            input.addEventListener(
                "click",
                function () {

                    // Select the current value immediately
                    // when the user clicks the field.
                    this.select();

                }
            );

        });
    }


    setupCustomerNameSelectAll();


    const observer = new MutationObserver(
        setupCustomerNameSelectAll
    );

    observer.observe(
        doc.body,
        {
            childList: true,
            subtree: true
        }
    );

    </script>
    """,
    height=0
)

# -------------------------------------------------------
# Amount State
# -------------------------------------------------------

if "amount_input" not in st.session_state:
    st.session_state["amount_input"] = ""

if "submit_from_enter" not in st.session_state:
    st.session_state["submit_from_enter"] = False

if "reset_amount" not in st.session_state:
    st.session_state["reset_amount"] = False


# Reset before Amount widget is created
if st.session_state["reset_amount"]:

    st.session_state["amount_input"] = ""
    st.session_state["reset_amount"] = False


# -------------------------------------------------------
# Enter Callback
# -------------------------------------------------------

def amount_entered():

    st.session_state[
        "submit_from_enter"
    ] = True

    st.session_state[
        "blur_amount"
    ] = True


# -------------------------------------------------------
# Amount
# -------------------------------------------------------

st.text_input(
    "Amount (Kg)",
    key="amount_input",
    on_change=amount_entered,
    placeholder="Enter amount"
)

# -------------------------------------------------------
# Remove Focus From Amount After Enter
# -------------------------------------------------------

if st.session_state.pop(
    "blur_amount",
    False
):

    components.html(
        """
        <script>
        const doc = window.parent.document;

        function blurAmount() {

            const labels = doc.querySelectorAll("label");

            for (const label of labels) {

                if (
                    label.innerText.trim() === "Amount (Kg)"
                ) {

                    const container = label.parentElement;

                    if (!container) {
                        return;
                    }

                    const input = container.querySelector(
                        "input"
                    );

                    if (!input) {
                        return;
                    }

                    input.blur();

                    return;
                }
            }
        }

        setTimeout(
            blurAmount,
            50
        );

        </script>
        """,
        height=0
    )

# -------------------------------------------------------
# Auto Focus Amount After Customer Selection
# -------------------------------------------------------

if st.session_state.pop(
    "focus_amount",
    False
):

    components.html(
        """
        <script>
        const doc = window.parent.document;

        function focusAmount() {

            const labels = doc.querySelectorAll("label");

            for (const label of labels) {

                if (
                    label.innerText.trim() === "Amount (Kg)"
                ) {

                    const container = label.parentElement;

                    if (!container) {
                        return;
                    }

                    const input = container.querySelector(
                        "input"
                    );

                    if (!input) {
                        return;
                    }

                    input.focus();

                    return;
                }
            }
        }

        setTimeout(
            focusAmount,
            100
        );
        </script>
        """,
        height=0
    )

# -------------------------------------------------------
# Convert Amount
# -------------------------------------------------------

try:

    amount = float(
        st.session_state["amount_input"]
    )

    amount_valid = amount >= 0

except (ValueError, TypeError):

    amount = 0.0
    amount_valid = False

# -------------------------------------------------------
# Submit Entry
# -------------------------------------------------------

submit_clicked = st.button(
    "Submit Entry",
    use_container_width=True,
    key="submit_entry_button"
)

submit_entry = (
    submit_clicked
    or st.session_state["submit_from_enter"]
)

# -------------------------------------------------------
# Adjust Existing Amount
# -------------------------------------------------------

with st.expander(
    "✏️ Adjust Existing Amount",
    expanded=False
):

    # Get current amount for selected customer/day
    try:

        adjustment_row = handler.find_customer_row(
            customer_code
        )

        adjustment_column = handler.find_date_column(
            selected_day
        )

        current_value = sheet.cell(
            row=adjustment_row,
            column=adjustment_column
        ).value

        try:
            current_amount = float(current_value)

        except (TypeError, ValueError):
            current_amount = 0.0

    except Exception:

        current_amount = 0.0


    st.markdown(
        f"**Current Amount:** "
        f"{current_amount:,.2f} Kg"
    )


    adjustment_type = st.radio(
        "Adjustment",
        [
            "➕ Add",
            "➖ Reduce"
        ],
        horizontal=True,
        key="adjustment_type"
    )


    adjustment_input = st.number_input(
        "Adjustment Amount (Kg)",
        min_value=0.0,
        step=0.5,
        value=0.0,
        key="adjustment_amount"
    )


    # Calculate preview
    if adjustment_type == "➕ Add":

        adjusted_amount = (
            current_amount
            + adjustment_input
        )

    else:

        adjusted_amount = (
            current_amount
            - adjustment_input
        )


    # Preview result
    if adjusted_amount >= 0:

        st.info(
            f"New Amount: "
            f"{adjusted_amount:,.2f} Kg"
        )

    else:

        st.error(
            "Adjustment would make the "
            "collection amount negative."
        )


    apply_adjustment = st.button(
        "Apply Adjustment",
        use_container_width=True,
        key="apply_adjustment_button"
    )

# -------------------------------------------------------
# Apply Adjustment
# -------------------------------------------------------

if apply_adjustment:

    if adjustment_input <= 0:

        st.error(
            "Please enter an adjustment amount "
            "greater than 0."
        )

    elif adjusted_amount < 0:

        st.error(
            "Reduction cannot be greater than "
            "the current amount."
        )

    else:

        try:

            old_value = current_amount


            # -------------------------------------------------------
            # Update Excel
            # -------------------------------------------------------

            row, column = handler.update(
                customer_code,
                selected_day,
                adjusted_amount
            )


            # -------------------------------------------------------
            # Update Database
            # -------------------------------------------------------

            save_collection(
                customer_code=customer_code,
                customer_name=customer_name,
                year=selected_year,
                month=selected_month,
                day=selected_day,
                amount=adjusted_amount
            )

            # Previous amount for audit
            try:
                previous_amount = float(old_value)
            except (TypeError, ValueError):
                previous_amount = None

            # -------------------------------------------------------
            # Determine Audit Action
            # -------------------------------------------------------

            if previous_amount is None:

                audit_action = "Added"

            else:

                audit_action = "Updated"

            # -------------------------------------------------------
            # Audit Record
            # -------------------------------------------------------

            add_audit_record(
                customer_code=customer_code,
                customer_name=customer_name,
                year=selected_year,
                month=selected_month,
                day=selected_day,
                old_amount=previous_amount,
                new_amount=adjusted_amount,
                username=st.session_state.get(
                    "username",
                    "Unknown"
                ),
                action=audit_action,
                excel_row=row,
                excel_column=column
            )


            # -------------------------------------------------------
            # Save Excel
            # -------------------------------------------------------

            handler.save(
                st.session_state["temp_file"]
            )


            # -------------------------------------------------------
            # Sync Workbook
            # -------------------------------------------------------

            sync_workbook(
                st.session_state["temp_file"],
                selected_year,
                selected_month
            )


            # -------------------------------------------------------
            # Update Highlight
            # -------------------------------------------------------

            st.session_state[
                "updated_cells"
            ].append(
                {
                    "row": row,
                    "column": column,
                    "year": selected_year,
                    "month": selected_month,
                    "day": selected_day,
                    "amount": adjusted_amount,
                    "time": datetime.now()
                }
            )


            st.session_state[
                "adjustment_success"
            ] = True

            st.rerun()


        except Exception as e:

            st.error(
                f"Adjustment failed: {e}"
            )

# -------------------------------------------------------
# Submit Success Message
# -------------------------------------------------------

if st.session_state.pop(
    "submit_success",
    False
):

    success_placeholder = st.empty()

    success_placeholder.success(
        "✅ Amount Added Successfully"
    )

    time.sleep(3)

    success_placeholder.empty()

# -------------------------------------------------------
# Adjustment Success Message
# -------------------------------------------------------

if st.session_state.pop(
    "adjustment_success",
    False
):

    adjustment_success_placeholder = st.empty()

    adjustment_success_placeholder.success(
        "✅ Amount Adjusted Successfully"
    )

    time.sleep(3)

    adjustment_success_placeholder.empty()


st.markdown("---")

# Find the selected Day column in DataFrame
day_column = None

for column in df.columns:

    if str(column) == str(selected_day):

        day_column = column
        break


# Calculate total collection for selected day
daily_total = 0.0

if day_column is not None:

    daily_values = pd.to_numeric(
        df[day_column],
        errors="coerce"
    )

    daily_total = daily_values.sum()


# -------------------------------------------------------
# Last Customer Entry
# -------------------------------------------------------

last_entry = get_last_collection(
    selected_year,
    selected_month
)


# -------------------------------------------------------
# Selected Date / Daily Total + Last Entry
# -------------------------------------------------------

if last_entry:

    last_customer_name = last_entry[0]
    last_amount = last_entry[1]
    last_year = last_entry[2]
    last_month = last_entry[3]
    last_day = last_entry[4]
    last_entry_time = last_entry[5]

    # ---------------------------------------------------
    # Convert database timestamp to Sri Lanka time
    # ---------------------------------------------------

    if last_entry_time is not None:

        # If PostgreSQL returned a timestamp without timezone,
        # treat the stored value as UTC
        if last_entry_time.tzinfo is None:

            last_entry_time = last_entry_time.replace(
                tzinfo=ZoneInfo("UTC")
            )

        # Convert UTC/database timezone to Sri Lanka time
        last_entry_time = last_entry_time.astimezone(
            ZoneInfo("Asia/Colombo")
        )

else:

    last_customer_name = "-"
    last_amount = None
    last_year = None
    last_month = None
    last_day = None
    last_entry_time = None


# -------------------------------------------------------
# SUMMARY ROW 1
# Selected Date | Last Customer + Amount
# -------------------------------------------------------

left_row1, right_row1 = st.columns([0.8, 1.2])

with left_row1:

    st.markdown(
        f"**📅 Date:** "
        f"{selected_day:02d} {selected_month} {selected_year}"
    )


with right_row1:

    if last_entry:

        st.markdown(
            f"""
<div style="text-align: right;">
<strong>👤 Last Customer:</strong> {last_customer_name}
&nbsp; | &nbsp;
<strong>🍃 Amount:</strong> {last_amount:,.2f} Kg
</div>
""",
            unsafe_allow_html=True
        )

    else:

        st.markdown(
            """
<div style="text-align: right;">
<strong>👤 Last Customer:</strong> -
&nbsp; | &nbsp;
<strong>🍃 Amount:</strong> -
</div>
""",
            unsafe_allow_html=True
        )


# -------------------------------------------------------
# SUMMARY ROW 2
# Daily Total | Last Entry Date + Time
# -------------------------------------------------------

left_row2, right_row2 = st.columns([0.8, 1.2])

with left_row2:

    st.markdown(
        f"**🍃 Daily Total Tea Collection:** "
        f"{daily_total:,.2f} Kg"
    )


with right_row2:

    if last_entry:

        st.markdown(
            f"""
<div style="text-align: right;">
<strong>📅 Date:</strong> {int(last_day):02d} {last_month} {last_year}
&nbsp; | &nbsp;
<strong>🕒 Time:</strong> {last_entry_time.strftime('%I:%M:%S %p')}
</div>
""",
            unsafe_allow_html=True
        )

    else:

        st.markdown(
            """
<div style="text-align: right;">
<strong>📅 Date:</strong> -
&nbsp; | &nbsp;
<strong>🕒 Time:</strong> -
</div>
""",
            unsafe_allow_html=True
        )

st.markdown("---")

if submit_entry:

    # Clear Enter flag immediately
    st.session_state["submit_from_enter"] = False

    if not amount_valid:

        st.error(
            "Please enter a valid Amount."
        )

    else:
        try:

            # -------------------------------------------------------
            # Get Existing Amount
            # -------------------------------------------------------

            old_value = sheet.cell(
                row=handler.find_customer_row(customer_code),
                column=handler.find_date_column(selected_day)
            ).value


            # -------------------------------------------------------
            # CASH Customer - Increment Existing Amount
            # -------------------------------------------------------

            if str(customer_name).strip().lower() == "cash":

                try:
                    existing_amount = float(old_value)

                except (TypeError, ValueError):
                    existing_amount = 0.0

                final_amount = existing_amount + amount

            else:

                # Normal customers replace existing amount
                final_amount = amount


            # -------------------------------------------------------
            # Update Excel
            # -------------------------------------------------------

            row, column = handler.update(
                customer_code,
                selected_day,
                final_amount
            )

            # -------------------------------------------------------
            # Save Collection Permanently
            # -------------------------------------------------------

            save_collection(
                customer_code=customer_code,
                customer_name=customer_name,
                year=selected_year,
                month=selected_month,
                day=selected_day,
                amount=final_amount
            )

            # -------------------------------------------------------
            # Previous amount for audit
            # -------------------------------------------------------

            try:

                previous_amount = float(old_value)

            except (TypeError, ValueError):

                previous_amount = None

            # -------------------------------------------------------
            # Determine Audit Action
            # -------------------------------------------------------

            # -------------------------------------------------------
            # Audit Action
            # -------------------------------------------------------

            if previous_amount is None:

                audit_action = "Added"

            else:

                audit_action = "Updated"

            # -------------------------------------------------------
            # Save Audit Record to SQLite
            # -------------------------------------------------------

            add_audit_record(
                customer_code=customer_code,
                customer_name=customer_name,
                year=selected_year,
                month=selected_month,
                day=selected_day,
                old_amount=previous_amount,
                new_amount=final_amount,
                username=st.session_state.get(
                    "username",
                    "Unknown"
                ),
                action=audit_action,
                excel_row=row,
                excel_column=column
            )

            handler.save(
                st.session_state["temp_file"]
            )

            # -------------------------------------------------------
            # Backup Updated Excel to Supabase Storage
            # -------------------------------------------------------

            sync_workbook(
                st.session_state["temp_file"],
                selected_year,
                selected_month
            )

            st.session_state["updated_cells"].append(
                {
                    "row": row,
                    "column": column,
                    "year": selected_year,
                    "month": selected_month,
                    "day": selected_day,
                    "amount": amount,
                    "time": datetime.now()
                }
            )

            # -------------------------------------------------------
            # Store Last Customer Entry
            # -------------------------------------------------------

            st.session_state["last_customer_name"] = customer_name

            st.session_state["last_entry_time"] = datetime.now(
                ZoneInfo("Asia/Colombo")
            )

         # Reset Amount on NEXT rerun
            st.session_state["reset_amount"] = True

            # Show success message after rerun
            st.session_state["submit_success"] = True

            st.rerun()

        except Exception as e:

            st.error(str(e))
# -------------------------------------------------------
# Current Excel Sheet
# -------------------------------------------------------

st.markdown(
    f"""
    <h3>
        Current Excel Sheet → [ {selected_year} - {selected_month} ]
    </h3>
    """,
    unsafe_allow_html=True
)

styled_df = (
    df.style
    .format(
        format_value,
        subset=df.columns[2:]
    )
    .apply(
        lambda data:
        highlight(
            data,
            st.session_state["updated_cells"]
        ),
        axis=None
    )
)


table_event = st.dataframe(
    styled_df,
    use_container_width=True,
    height=500,
    on_select="rerun",
    selection_mode="single-cell",
    key="tea_collection_table"
)

# -------------------------------------------------------
# Selected Table Cell
# -------------------------------------------------------

selected_cell = None
selected_value = None
selected_row_index = None
selected_column = None

remove_enabled = False


if table_event.selection.cells:

    selected_cell = (
        table_event.selection.cells[0]
    )

    # Streamlit returns:
    # (row_position, column_name)
    selected_row_index = selected_cell[0]
    selected_column = selected_cell[1]


    # ---------------------------------------------------
    # Convert Day column to integer
    # ---------------------------------------------------

    if selected_column not in [
        "Customer Code",
        "Customer Name"
    ]:

        try:

            selected_column = int(
                selected_column
            )

        except (ValueError, TypeError):

            selected_column = None


    # ---------------------------------------------------
    # Get Selected Value
    # ---------------------------------------------------

    if (
        selected_column is not None
        and selected_column in df.columns
    ):

        selected_value = df.iloc[
            selected_row_index
        ][selected_column]


        # -----------------------------------------------
        # Only Day columns containing a value
        # can be removed
        # -----------------------------------------------

        if selected_column not in [
            "Customer Code",
            "Customer Name"
        ]:

            if (
                selected_value is not None
                and not pd.isna(selected_value)
                and str(selected_value).strip() not in ["", "-"]
            ):

                remove_enabled = True

if remove_enabled:

    remove_customer_code = df.iloc[
        selected_row_index
    ]["Customer Code"]

    remove_customer_name = df.iloc[
        selected_row_index
    ]["Customer Name"]

    st.info(
        f"Selected Entry: "
        f"{remove_customer_code} - "
        f"{remove_customer_name} | "
        f"Day {selected_column} | "
        f"{selected_value} Kg"
    )

remove_clicked = st.button(
    "🗑️ Remove Entry",
    disabled=not remove_enabled,
    use_container_width=True,
    key="remove_selected_entry"
)


# -------------------------------------------------------
# Remove Entry Success Message
# -------------------------------------------------------

if st.session_state.pop(
    "remove_success",
    False
):

    st.success(
        "✅ Entry removed successfully."
    )

if remove_clicked:

    st.session_state[
        "remove_entry_data"
    ] = {
        "row_index": selected_row_index,
        "column": selected_column,
        "value": selected_value,
        "customer_code": remove_customer_code,
        "customer_name": remove_customer_name
    }

    st.session_state[
        "show_remove_confirmation"
    ] = True

    st.rerun()

# -------------------------------------------------------
# Remove Entry Confirmation
# -------------------------------------------------------

if st.session_state.get(
    "show_remove_confirmation",
    False
):

    remove_data = st.session_state.get(
        "remove_entry_data"
    )

    if remove_data:

        st.warning(
            f"⚠️ Remove "
            f"{remove_data['value']} Kg for "
            f"{remove_data['customer_name']} "
            f"(Customer Code: "
            f"{remove_data['customer_code']}) "
            f"on Day {remove_data['column']}?"
        )

        confirm_col1, confirm_col2 = st.columns(2)


        # ---------------------------------------------------
        # Cancel Button
        # ---------------------------------------------------

        with confirm_col1:

            cancel_remove = st.button(
                "Cancel",
                use_container_width=True,
                key="cancel_remove_entry"
            )


        # ---------------------------------------------------
        # Confirm Button
        # ---------------------------------------------------

        with confirm_col2:

            confirm_remove = st.button(
                "🗑️ Confirm Remove",
                type="primary",
                use_container_width=True,
                key="confirm_remove_entry"
            )


        # ---------------------------------------------------
        # Cancel Remove
        # ---------------------------------------------------

        if cancel_remove:

            st.session_state[
                "show_remove_confirmation"
            ] = False

            st.session_state.pop(
                "remove_entry_data",
                None
            )

            st.rerun()


        # ---------------------------------------------------
        # Confirm Remove
        # ---------------------------------------------------

        if confirm_remove:

            with st.spinner(
                "Removing entry and syncing workbook..."
            ):

                try:

                    # -------------------------------------------
                    # Get selected entry information
                    # -------------------------------------------

                    remove_customer_code = (
                        remove_data["customer_code"]
                    )

                    remove_customer_name = (
                        remove_data["customer_name"]
                    )

                    remove_day = int(
                        remove_data["column"]
                    )

                    old_amount = remove_data[
                        "value"
                    ]


                    # -------------------------------------------
                    # Find corresponding Excel cell
                    # -------------------------------------------

                    excel_row = (
                        handler.find_customer_row(
                            remove_customer_code
                        )
                    )

                    excel_column = (
                        handler.find_date_column(
                            remove_day
                        )
                    )


                    # -------------------------------------------
                    # Clear Excel cell
                    # -------------------------------------------

                    sheet.cell(
                        row=excel_row,
                        column=excel_column
                    ).value = None


                    # -------------------------------------------
                    # Delete PostgreSQL record
                    # -------------------------------------------

                    delete_collection(
                        customer_code=remove_customer_code,
                        year=selected_year,
                        month=selected_month,
                        day=remove_day
                    )

                    # -------------------------------------------------------
                    # Previous amount for audit
                    # -------------------------------------------------------

                    try:

                        previous_amount = float(old_amount)

                    except (TypeError, ValueError):

                        previous_amount = None


                    # -------------------------------------------
                    # Add deletion to Audit
                    # -------------------------------------------

                    add_audit_record(
                        customer_code=remove_customer_code,
                        customer_name=remove_customer_name,
                        year=selected_year,
                        month=selected_month,
                        day=remove_day,
                        old_amount=previous_amount,
                        new_amount=0,
                        username=st.session_state.get(
                            "username",
                            "Unknown"
                        ),
                        action="Deleted",
                        excel_row=excel_row,
                        excel_column=excel_column
                    )


                    # -------------------------------------------
                    # Save local workbook
                    # -------------------------------------------

                    handler.save(
                        st.session_state["temp_file"]
                    )


                    # -------------------------------------------
                    # Sync updated workbook to Supabase
                    # -------------------------------------------

                    sync_workbook(
                        st.session_state["temp_file"],
                        selected_year,
                        selected_month
                    )


                    # -------------------------------------------
                    # Remove highlight if applicable
                    # -------------------------------------------

                    st.session_state[
                        "updated_cells"
                    ] = [
                        item
                        for item in st.session_state[
                            "updated_cells"
                        ]
                        if not (
                            item.get("row") == excel_row
                            and
                            item.get("column") == excel_column
                        )
                    ]


                    # -------------------------------------------
                    # Close confirmation
                    # -------------------------------------------

                    st.session_state[
                        "show_remove_confirmation"
                    ] = False

                    st.session_state.pop(
                        "remove_entry_data",
                        None
                    )


                    # Show success message after rerun
                    st.session_state["remove_success"] = True

                    st.rerun()


                except Exception as e:

                    st.error(
                        f"Unable to remove entry: {e}"
                    )

# -------------------------------------------------------
# Navigation Drawer
# -------------------------------------------------------

with st.sidebar:

    st.write(
    f"👤 {st.session_state.get('username', '')}"
    )

    if st.button(
        "🚪 Logout",
        use_container_width=True,
        key="logout_button"
    ):

        logout()

        st.rerun()

    st.title("🍃 Menu")
    
    st.markdown("---")

    # ---------------------------------------------------
    # Download Workbook
    # ---------------------------------------------------

    with open(
        st.session_state["temp_file"],
        "rb"
    ) as file:

        st.download_button(
            label="📥 Download Workbook",

            data=file.read(),

            file_name=(
                f"Tea_Collection_"
                f"{selected_year}_"
                f"{selected_month}.xlsx"
            ),

            mime=(
                "application/"
                "vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),

            use_container_width=True
        )

    st.markdown("---")

    # ---------------------------------------------------
    # Audit Panel
    # ---------------------------------------------------

    if st.button(
        "🔐 Audit Panel",
        use_container_width=True,
        key="audit_panel_button"
    ):

        st.session_state["show_audit_panel"] = (
            not st.session_state.get(
                "show_audit_panel",
                False
            )
        )


    if st.session_state.get(
        "show_audit_panel",
        False
    ):    

        if not st.session_state.get(
            "audit_authenticated",
            False
        ):

            st.subheader("🔐 Audit Access")

            audit_password = st.text_input(
                "Password",
                type="password",
                key="audit_password"
            )

            if st.button(
                "Unlock",
                use_container_width=True,
                key="unlock_audit_button"
            ):

                if verify_audit_password(
                    audit_password
                ):

                    st.session_state[
                        "audit_authenticated"
                    ] = True

                    st.rerun()

                else:

                    st.error(
                        "Incorrect password."
                    )

        else:

            # ---------------------------------------------------
            # Lock Audit Panel
            # ---------------------------------------------------

            if st.button(
                "🔒 Lock Audit Panel",
                use_container_width=True,
                key="lock_audit_button"
            ):

                st.session_state["audit_authenticated"] = False
                st.session_state["show_audit_panel"] = False

                st.rerun()

            st.subheader("📋 Audit History")

            audit_records = get_audit_records()

            # ---------------------------------------------------
            # Download Audit History
            # ---------------------------------------------------

            if audit_records:

                audit_download_data = []

                for record in audit_records:

                    (
                        record_id,
                        customer_code,
                        customer_name,
                        year,
                        month,
                        day,
                        old_amount,
                        new_amount,
                        username,
                        action,
                        updated_at
                    ) = record

                    updated_time = datetime.fromisoformat(
                        updated_at
                    )

                    audit_download_data.append({
                        "Audit ID": record_id,
                        "Customer Code": customer_code,
                        "Customer Name": customer_name,
                        "Year": year,
                        "Month": month,
                        "Day": day,
                        "Collection Date": (
                            f"{int(day):02d} "
                            f"{month} "
                            f"{year}"
                        ),
                        "Action": action,
                        "Old Amount (Kg)": old_amount,
                        "New Amount (Kg)": new_amount,
                        "User": username,
                        "Updated": updated_time.strftime(
                            "%d-%m-%Y %I:%M:%S %p"
                        )
                    })


                audit_df = pd.DataFrame(
                    audit_download_data
                )

                # ---------------------------------------------------
                # Sort records from oldest to newest
                # ---------------------------------------------------

                audit_df = audit_df.sort_values(
                    by="Audit ID",
                    ascending=True
                ).reset_index(drop=True)


                # ---------------------------------------------------
                # Create sequential Audit ID
                # ---------------------------------------------------

                audit_df["Audit ID"] = range(
                    1,
                    len(audit_df) + 1
                )

                # ---------------------------------------------------
                # Create Professional Excel Report
                # ---------------------------------------------------

                output = io.BytesIO()

                with pd.ExcelWriter(
                    output,
                    engine="openpyxl"
                ) as writer:

                    audit_df.to_excel(
                        writer,
                        index=False,
                        sheet_name="Audit History"
                    )

                    workbook = writer.book

                    worksheet = writer.sheets[
                        "Audit History"
                    ]


                    # ------------------------------------------------
                    # Header Formatting
                    # ------------------------------------------------

                    from openpyxl.styles import (
                        Font,
                        PatternFill,
                        Alignment,
                        Border,
                        Side
                    )

                    header_fill = PatternFill(
                        fill_type="solid",
                        fgColor="1F2937"
                    )

                    header_font = Font(
                        color="FFFFFF",
                        bold=True
                    )

                    header_alignment = Alignment(
                        horizontal="center",
                        vertical="center"
                    )


                    for cell in worksheet[1]:

                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = header_alignment


                    # ------------------------------------------------
                    # Freeze Header
                    # ------------------------------------------------

                    worksheet.freeze_panes = "A2"


                    # ------------------------------------------------
                    # Enable Filter
                    # ------------------------------------------------

                    worksheet.auto_filter.ref = (
                        worksheet.dimensions
                    )


                    # ------------------------------------------------
                    # Column Widths
                    # ------------------------------------------------

                    column_widths = {

                        "A": 10,   # Audit ID
                        "B": 18,   # Customer Code
                        "C": 28,   # Customer Name
                        "D": 10,   # Year
                        "E": 15,   # Month
                        "F": 10,   # Day
                        "G": 22,   # Collection Date
                        "H": 14,   # Action
                        "I": 18,   # Old Amount
                        "J": 18,   # New Amount
                        "K": 18,   # User
                        "L": 25    # Updated
                    }


                    for column, width in column_widths.items():

                        worksheet.column_dimensions[
                            column
                        ].width = width


                    # ------------------------------------------------
                    # Cell Alignment
                    # ------------------------------------------------

                    for row in worksheet.iter_rows(
                        min_row=2
                    ):

                        for cell in row:

                            cell.alignment = Alignment(
                                vertical="center"
                            )


                    # ------------------------------------------------
                    # Action Formatting
                    # ------------------------------------------------

                    for row in range(
                        2,
                        worksheet.max_row + 1
                    ):

                        action_cell = worksheet.cell(
                            row=row,
                            column=8
                        )

                        action_value = (
                            action_cell.value
                        )


                        if action_value == "Added":

                            action_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor="DCFCE7"
                            )

                            action_cell.font = Font(
                                color="166534",
                                bold=True
                            )


                        elif action_value == "Updated":

                            action_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor="DBEAFE"
                            )

                            action_cell.font = Font(
                                color="1D4ED8",
                                bold=True
                            )


                        elif action_value == "Deleted":

                            action_cell.fill = PatternFill(
                                fill_type="solid",
                                fgColor="FEE2E2"
                            )

                            action_cell.font = Font(
                                color="B91C1C",
                                bold=True
                            )


                    # ------------------------------------------------
                    # Amount Formatting
                    # ------------------------------------------------

                    for row in range(
                        2,
                        worksheet.max_row + 1
                    ):

                        worksheet.cell(
                            row=row,
                            column=9
                        ).number_format = '0.00'

                        worksheet.cell(
                            row=row,
                            column=10
                        ).number_format = '0.00'


                    # ------------------------------------------------
                    # Add Excel Table
                    # ------------------------------------------------

                    from openpyxl.worksheet.table import (
                        Table,
                        TableStyleInfo
                    )

                    table_reference = (
                        f"A1:L{worksheet.max_row}"
                    )

                    audit_table = Table(
                        displayName="AuditHistoryTable",
                        ref=table_reference
                    )

                    table_style = TableStyleInfo(
                        name="TableStyleMedium2",
                        showFirstColumn=False,
                        showLastColumn=False,
                        showRowStripes=True,
                        showColumnStripes=False
                    )

                    audit_table.tableStyleInfo = (
                        table_style
                    )

                    worksheet.add_table(
                        audit_table
                    )


                    # ------------------------------------------------
                    # Row Height
                    # ------------------------------------------------

                    worksheet.row_dimensions[1].height = 25


                output.seek(0)


                # ---------------------------------------------------
                # Download Button
                # ---------------------------------------------------

                st.download_button(

                    label="📥 Download Audit History",

                    data=output,

                    file_name=(
                        "Tea_Collection_Audit_History.xlsx"
                    ),

                    mime=(
                        "application/"
                        "vnd.openxmlformats-officedocument."
                        "spreadsheetml.sheet"
                    ),

                    use_container_width=True,

                    key="download_audit_history"
                )

                st.markdown("---")

            if not audit_records:

                st.info(
                    "No audit records available."
                )

            else:

                for record in audit_records[:5]:

                    (
                        record_id,
                        customer_code,
                        customer_name,
                        year,
                        month,
                        day,
                        old_amount,
                        new_amount,
                        username,
                        action,
                        updated_at
                    ) = record

                    # ---------------------------------------------------
                    # Format Date / Time
                    # ---------------------------------------------------

                    updated_time = datetime.fromisoformat(
                        updated_at
                    )

                    formatted_date = (
                        f"{int(day):02d} {month} {year}"
                    )

                    formatted_time = updated_time.strftime(
                        "%d %b %Y • %I:%M:%S %p"
                    )

                    # ---------------------------------------------------
                    # Action Style
                    # ---------------------------------------------------

                    if action == "Added":

                        badge_background = "#DCFCE7"
                        badge_text = "#166534"
                        badge_icon = "🟢"

                    elif action == "Updated":

                        badge_background = "#DBEAFE"
                        badge_text = "#1D4ED8"
                        badge_icon = "🔵"

                    else:

                        badge_background = "#FEE2E2"
                        badge_text = "#B91C1C"
                        badge_icon = "🔴"

                    # ---------------------------------------------------
                    # Amount Section
                    # ---------------------------------------------------

                    if action == "Added":

                        amount_html = f"""
                        <div style="
                            padding:12px 0;
                            border-top:1px solid #374151;
                        ">
                            <div style="
                                font-size:12px;
                                color:#6B7280;
                                margin-bottom:3px;
                            ">
                                🍃 New Amount
                            </div>

                            <div style="
                                font-size:18px;
                                font-weight:600;
                            ">
                                {new_amount:g} Kg
                            </div>
                        </div>
                        """

                    elif action == "Updated":

                        old_amount_display = (
                            f"{old_amount:g} Kg"
                            if old_amount is not None
                            else "-"
                        )

                        new_amount_display = (
                            f"{new_amount:g} Kg"
                            if new_amount is not None
                            else "-"
                        )

                        amount_html = f"""
                        <div style="
                            display:grid;
                            grid-template-columns:1fr 1fr;
                            gap:20px;
                            padding:12px 0;
                            border-top:1px solid #374151;
                        ">

                            <div>

                                <div style="
                                    font-size:12px;
                                    color:#6B7280;
                                    margin-bottom:3px;
                                ">
                                    🍃 Old Amount
                                </div>

                                <div style="
                                    font-size:18px;
                                    font-weight:600;
                                    color:#F9FAFB;
                                ">
                                    {old_amount_display}
                                </div>

                            </div>


                            <div>

                                <div style="
                                    font-size:12px;
                                    color:#6B7280;
                                    margin-bottom:3px;
                                ">
                                    🍃 New Amount
                                </div>

                                <div style="
                                    font-size:18px;
                                    font-weight:600;
                                    color:#F9FAFB;
                                ">
                                    {new_amount_display}
                                </div>

                            </div>

                        </div>
                        """
                    else:

                        removed_amount = (
                            old_amount
                            if old_amount is not None
                            else 0
                        )

                        amount_html = f"""
                        <div style="
                            padding:12px 0;
                            border-top:1px solid #374151;
                        ">
                            <div style="
                                font-size:12px;
                                color:#6B7280;
                                margin-bottom:3px;
                            ">
                                🍃 Removed Amount
                            </div>

                            <div style="
                                font-size:18px;
                                font-weight:600;
                                color:#F9FAFB;
                            ">
                                {removed_amount:g} Kg
                            </div>
                        </div>
                        """

                    # ---------------------------------------------------
                    # Audit Card
                    # ---------------------------------------------------

                    audit_card_html = textwrap.dedent(
                        f"""
                        <div style="
                            border:1px solid #374151;
                            border-radius:12px;
                            padding:16px;
                            margin-bottom:14px;
                            background:#1F2937;
                            box-shadow:0 2px 6px rgba(0,0,0,0.25);
                            color:#F9FAFB;
                        ">

                            <!-- Header -->

                            <div style="
                                display:flex;
                                justify-content:space-between;
                                align-items:center;
                                margin-bottom:14px;
                            ">

                                <div style="
                                    font-size:12px;
                                    font-weight:700;
                                    color:#6B7280;
                                ">
                                    AUDIT #{record_id}
                                </div>

                                <div style="
                                    display:inline-block;
                                    padding:5px 10px;
                                    border-radius:20px;
                                    background:{badge_background};
                                    color:{badge_text};
                                    font-size:12px;
                                    font-weight:700;
                                ">
                                    {badge_icon} {action.upper()}
                                </div>

                            </div>

                            <!-- Customer Code -->

                            <div style="
                                margin-bottom:10px;
                            ">

                                <div style="
                                    font-size:12px;
                                    color:#6B7280;
                                    margin-bottom:3px;
                                ">
                                    👤 Customer Code
                                </div>

                                <div style="
                                    font-size:16px;
                                    font-weight:600;
                                    color:#F9FAFB;
                                ">
                                    {customer_code}
                                </div>

                            </div>

                            <!-- Customer Name -->

                            <div style="
                                margin-bottom:10px;
                            ">

                                <div style="
                                    font-size:12px;
                                    color:#6B7280;
                                    margin-bottom:3px;
                                ">
                                    👤 Customer Name
                                </div>

                                <div style="
                                    font-size:16px;
                                    font-weight:600;
                                    color:#F9FAFB;
                                ">
                                    {customer_name}
                                </div>

                            </div>

                            <!-- Collection Date -->

                            <div style="
                                margin-bottom:4px;
                            ">

                                <div style="
                                    font-size:12px;
                                    color:#6B7280;
                                    margin-bottom:3px;
                                ">
                                    📅 Collection Date
                                </div>

                                <div style="
                                    font-size:14px;
                                    font-weight:500;
                                    color:#F9FAFB;
                                ">
                                    {formatted_date}
                                </div>

                            </div>

                            <!-- Amount -->

                            {amount_html}

                            <!-- Footer -->

                            <div style="
                                display:grid;
                                grid-template-columns:1fr 1fr;
                                gap:20px;
                                border-top:1px solid #374151;
                                padding-top:12px;
                                margin-top:4px;
                            ">

                                <div>

                                    <div style="
                                        font-size:12px;
                                        color:#6B7280;
                                        margin-bottom:3px;
                                    ">
                                        👤 User
                                    </div>

                                    <div style="
                                        font-size:14px;
                                        font-weight:600;
                                        color:#F9FAFB;
                                    ">
                                        {username}
                                    </div>

                                </div>

                                <div>

                                    <div style="
                                        font-size:12px;
                                        color:#6B7280;
                                        margin-bottom:3px;
                                    ">
                                        🕒 Updated
                                    </div>

                                    <div style="
                                        font-size:14px;
                                        font-weight:600;
                                        color:#F9FAFB;
                                    ">
                                        {formatted_time}
                                    </div>

                                </div>

                            </div>

                        </div>
                        """
                    )

                    st.html(audit_card_html)
                            